import re
import math
from collections import defaultdict
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

import pdfplumber
import openpyxl
import streamlit as st


# ---------- helpers ----------
def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


def decimal_2(x) -> Decimal | None:
    if x is None:
        return None
    if isinstance(x, Decimal):
        return x.quantize(Decimal("0.01"))
    if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x)):
        return Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    s = str(x).strip()
    if "," in s and "." not in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    if s.startswith("."):
        s = "0" + s
    return Decimal(s).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def get_col_idx_by_header(ws, header_contains: str) -> int:
    target = norm(header_contains)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        if target in norm(str(v)):
            return c
    raise ValueError(f"No encontré la columna cuyo header contenga: {header_contains}")


def get_col_idx_prefer_exact(ws, exact_headers: list[str], contains_fallback: str | None = None) -> int:
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        header_map[norm(str(v))] = c

    for h in exact_headers:
        key = norm(h)
        if key in header_map:
            return header_map[key]

    if contains_fallback:
        target = norm(contains_fallback)
        for k, c in header_map.items():
            if target in k:
                return c

    raise ValueError(f"No encontré columna. exact={exact_headers} fallback={contains_fallback}")


def parse_statement_year_from_pdf_bytes(pdf_bytes: bytes) -> int | None:
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        t = pdf.pages[0].extract_text() or ""
    m = re.search(r"DEL\s*\d{2}/\d{2}/(\d{4})\s*AL\s*\d{2}/\d{2}/(\d{4})", t)
    return int(m.group(1)) if m else None


DATE_RE = re.compile(r"^\s*(\d{2})-(\d{2})\b")
AMT_TOKEN_RE = re.compile(r"(?:\d{1,3}(?:,\d{3})*\.\d{2}|\.\d{2})-?")


def extract_bcp_negative_transactions_from_bytes(pdf_bytes: bytes, default_year: int) -> list[dict]:
    year = parse_statement_year_from_pdf_bytes(pdf_bytes) or default_year
    tx = []

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for pnum, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line in text.splitlines():
                md = DATE_RE.match(line)
                if not md:
                    continue

                dd, mm = int(md.group(1)), int(md.group(2))

                # Busca montos con decimales y opcional '-'
                amounts = AMT_TOKEN_RE.findall(line)
                if not amounts:
                    continue

                # ✅ En tu PDF hay líneas donde solo aparece el cargo negativo (ej 110.23-)
                # así que tomamos el ÚLTIMO monto negativo en la línea (si existe).:contentReference[oaicite:3]{index=3}
                neg_tok = None
                for a in reversed(amounts):
                    if a.endswith("-"):
                        neg_tok = a
                        break
                if not neg_tok:
                    continue

                amt = -decimal_2(neg_tok[:-1])
                num_op = extract_num_op_from_line(line)

                tx.append({
                    "page": pnum,
                    "fecha_proc": date(year, mm, dd),
                    "num_op": num_op,
                    "amount": amt,       # negativo
                    "line": line.strip(),
                })

    return tx


TIME_RE = re.compile(r"\b\d{2}:\d{2}\b")

def extract_num_op_from_line(line: str) -> str:
    """
    BCP: NUM OP suele estar inmediatamente antes de la HORA (HH:MM).
    Ej: '... 408357 07:53 ... 110.23-':contentReference[oaicite:2]{index=2}
    También hay líneas tipo ITF: '... INT - 0909 .30-' (sin hora).
    """
    tokens = line.strip().split()

    # Caso 1: hay hora -> token anterior es NUM OP
    for i, tok in enumerate(tokens):
        if TIME_RE.fullmatch(tok):
            if i > 0 and re.fullmatch(r"\d{3,}", tokens[i - 1] or ""):
                return tokens[i - 1]
            break

    # Caso 2: líneas sin hora, tipo "... - 0909 .30-"
    m = re.search(r"\b-\s*(\d{3,})\b", line)
    if m:
        return m.group(1)

    return ""

def run_match(pdf_bytes: bytes, sire_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(sire_bytes), data_only=True)
    ws = wb.active

    col_fecha = get_col_idx_by_header(ws, "Fecha de emisión")
    col_total = get_col_idx_by_header(ws, "Total CP")
    col_docid = get_col_idx_by_header(ws, "Nro Doc Identidad")

    # 👇 usar SOLO la 2da columna (la correcta)
    col_nombre = get_col_idx_prefer_exact(
        ws,
        ["Apellidos Nombres/ Razón  Social", "Apellidos Nombres/ Razón Social", "Apellidos Nombres/Razón Social"],
        contains_fallback="apellidos nombres/ razon",
    )

    # default_year desde SIRE
    years = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_fecha).value
        if isinstance(v, datetime):
            years.append(v.year)
    default_year = min(years) if years else datetime.now().year

    bcp_tx = extract_bcp_negative_transactions_from_bytes(pdf_bytes, default_year)

    bcp_index = defaultdict(list)
    for i, t in enumerate(bcp_tx):
        bcp_index[(abs(t["amount"]), t["fecha_proc"])].append(i)

    used_bcp = set()
    matches = []

    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, col_fecha).value
        total = ws.cell(r, col_total).value
        if fecha is None or total is None:
            continue

        sire_date = fecha.date() if isinstance(fecha, datetime) else fecha
        amt = decimal_2(total)
        if amt is None:
            continue

        candidates = []
        for d in (sire_date, sire_date + timedelta(days=1)):
            for idx in bcp_index.get((amt, d), []):
                if idx not in used_bcp:
                    candidates.append((d, idx))

        if not candidates:
            continue

        candidates.sort(key=lambda x: 0 if x[0] == sire_date else 1)
        chosen_d, idx = candidates[0]
        used_bcp.add(idx)

        docid = ws.cell(r, col_docid).value
        if isinstance(docid, (int, float)) and not (isinstance(docid, float) and math.isnan(docid)):
            docid_str = str(int(docid))
        else:
            docid_str = (str(docid).strip() if docid is not None else "")

        nombre = ws.cell(r, col_nombre).value

        matches.append(
            {
                "Fecha emision": sire_date,  # sigue siendo la del Excel SIRE
                "Nro Doc Identidad": docid_str,
                "Apellidos Nombre / Razon Social": nombre,

                "FECHA PROC": bcp_tx[idx]["fecha_proc"],
                "NUM OP": bcp_tx[idx].get("num_op", ""),

                "Total CP": float(amt),
            }
        )


    return matches


def build_xlsx_bytes(matches: list[dict]) -> bytes:
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "matches"

    headers = ["Fecha emision", "Nro Doc Identidad", "Apellidos Nombre / Razon Social", "FECHA PROC", "NUM OP", "Total CP"]
    out_ws.append(headers)

    for m in matches:
        out_ws.append([
            m["Fecha emision"],
            m["Nro Doc Identidad"],
            m["Apellidos Nombre / Razon Social"],
            m["FECHA PROC"],
            m["NUM OP"],
            m["Total CP"],
        ])

    # formatos
    out_ws.column_dimensions["A"].width = 14
    out_ws.column_dimensions["B"].width = 18
    out_ws.column_dimensions["C"].width = 55
    out_ws.column_dimensions["D"].width = 14
    out_ws.column_dimensions["E"].width = 14
    out_ws.column_dimensions["F"].width = 12

    for cell in out_ws["A"][1:]:
        cell.number_format = "dd/mm/yyyy"
    for cell in out_ws["D"][1:]:
        cell.number_format = "dd/mm/yyyy"
    for cell in out_ws["F"][1:]:
        cell.number_format = "0.00"


    bio = BytesIO()
    out_wb.save(bio)
    return bio.getvalue()


def build_txt_bytes(matches: list[dict]) -> bytes:
    headers = ["Fecha emision", "Nro Doc Identidad", "Apellidos Nombre / Razon Social", "FECHA PROC", "NUM OP", "Total CP"]
    lines = ["|".join(headers)]
    for m in matches:
        lines.append(
            f"{m['Fecha emision'].strftime('%d/%m/%Y')}|{m['Nro Doc Identidad']}|{m['Apellidos Nombre / Razon Social']}|"
            f"{m['FECHA PROC'].strftime('%d/%m/%Y')}|{m['NUM OP']}|{m['Total CP']:.2f}"
        )
    
    return ("\n".join(lines) + "\n").encode("utf-8")


# ---------- UI ----------
st.set_page_config(page_title="BCP vs SIRE Matcher", layout="wide")
st.title("Comparación BCP (PDF) vs SIRE (XLSX)")

# ✅ Estado persistente
if "matches" not in st.session_state:
    st.session_state.matches = None
if "xlsx_bytes" not in st.session_state:
    st.session_state.xlsx_bytes = None
if "txt_bytes" not in st.session_state:
    st.session_state.txt_bytes = None
if "view_rows" not in st.session_state:
    st.session_state.view_rows = None
if "last_ok" not in st.session_state:
    st.session_state.last_ok = False

col1, col2 = st.columns(2)
with col1:
    pdf_file = st.file_uploader("Sube tu BCP.pdf", type=["pdf"], key="pdf_uploader")
with col2:
    xlsx_file = st.file_uploader("Sube tu SIRE.xlsx", type=["xlsx"], key="xlsx_uploader")

run = st.button("Procesar", type="primary", disabled=not (pdf_file and xlsx_file), key="btn_process")

# ✅ Procesar y guardar en session_state
if run:
    try:
        with st.spinner("Procesando..."):
            pdf_bytes = pdf_file.read()
            sire_bytes = xlsx_file.read()

            matches = run_match(pdf_bytes, sire_bytes)

            view_rows = [
                {
                    "Fecha emision": m["Fecha emision"].strftime("%d/%m/%Y"),
                    "Nro Doc Identidad": m["Nro Doc Identidad"],
                    "Apellidos Nombre / Razon Social": m["Apellidos Nombre / Razon Social"],
                    "FECHA PROC": m["FECHA PROC"].strftime("%d/%m/%Y"),
                    "NUM OP": m["NUM OP"],
                    "Total CP": f"{m['Total CP']:.2f}",
                }
                for m in matches
            ]

            st.session_state.matches = matches
            st.session_state.view_rows = view_rows
            st.session_state.xlsx_bytes = build_xlsx_bytes(matches)
            st.session_state.txt_bytes = build_txt_bytes(matches)
            st.session_state.last_ok = True

    except Exception as e:
        st.session_state.last_ok = False
        st.error(f"Error: {e}")

# ✅ Mostrar resultados SIEMPRE que existan (aunque hayas descargado)
if st.session_state.matches is not None:
    if st.session_state.last_ok:
        st.success(f"Listo. Coincidencias encontradas: {len(st.session_state.matches)}")

    st.dataframe(st.session_state.view_rows, use_container_width=True, height=420)

    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "⬇️ Descargar XLSX",
            data=st.session_state.xlsx_bytes,
            file_name="SIRE_BCP_matches.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_xlsx",
        )
    with d2:
        st.download_button(
            "⬇️ Descargar TXT",
            data=st.session_state.txt_bytes,
            file_name="SIRE_BCP_matches.txt",
            mime="text/plain",
            use_container_width=True,
            key="dl_txt",
        )

    # opcional: botón para limpiar
    if st.button("Limpiar resultado", key="btn_clear"):
        st.session_state.matches = None
        st.session_state.view_rows = None
        st.session_state.xlsx_bytes = None
        st.session_state.txt_bytes = None
        st.session_state.last_ok = False
